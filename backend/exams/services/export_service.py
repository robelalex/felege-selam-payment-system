"""
Excel export for the homeroom "Check Result and Award" screen's overview
table — Term/Semester columns + Year Average + Rank, exactly what's
already shown on screen (TeacherClassResults.js overview mode), just as a
downloadable .xlsx so a teacher can save/print/sort it outside the app.

Deliberately reuses the exact same computed data the on-screen table and
the cumulative report card both already use (report_cards.services.
cumulative_service) — this must never show a different number than the
screen or the PDF, so it never recomputes anything itself.
"""
import io
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from django.http import HttpResponse

HEADER_FILL = PatternFill(start_color='0F4C81', end_color='0F4C81', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')


def _as_float(value):
    """
    Normalizes a period/overall average to a plain float before writing it
    to a cell. Without this, Django's DecimalField values (per-term/per-
    semester averages) fail the `isinstance(value, (int, float))` check
    below and silently skip the % number format, while the one plain-float
    value (overall_average) gets it — exactly the inconsistent formatting
    bug caught in visual verification (quarters showed "73" but Year
    Average showed "80.5%" on the same row). Casting everything to float
    up front makes every percentage column behave identically.
    """
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def build_class_results_workbook(*, title, subtitle, period_names, results, period_key, average_key, rank_key, rank_total_key):
    """
    Builds the workbook and returns raw bytes.

    title / subtitle: strings shown in the first two rows (school + class,
        e.g. "Jimma Excellence Academy" / "Grade 10 F — Year-End Results (Quarter view)").
    period_names: list of {'id', 'name'} — one column per period (term or semester).
    results: list of dicts, one per student, each already containing the
        period-level breakdown (a list under `period_key`, each entry with
        'average' keyed by the matching id field), the overall figure under
        `average_key`, and rank under `rank_key` / `rank_total_key`. This is
        exactly the same `results` list the JSON endpoints already return —
        no separate computation path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Class Results'

    n_periods = len(period_names)
    total_cols = 4 + n_periods + 3  # Rank, Out Of, Student, Student ID, [periods...], Average, Grade, Result

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color='0F4C81')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=10, color='6B7280')

    header_row = 4
    headers = ['Rank', 'Out Of', 'Student Name', 'Student ID'] + [p['name'] for p in period_names] + ['Year Average', 'Grade', 'Result']
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row_i = header_row + 1
    for r in results:
        periods_by_id = {p['id'] if 'id' in p else p.get('term_id', p.get('semester_id')): p.get('average')
                          for p in r.get(period_key, [])}
        # The two JSON shapes differ slightly in key naming (term_id/semester_id
        # vs plain id) depending on which endpoint built `results` — handle
        # both without assuming one.
        row_values = [
            r.get(rank_key),
            r.get(rank_total_key),
            r.get('student_name', ''),
            r.get('student_id_display', ''),
        ]
        for p in period_names:
            entry = next((x for x in r.get(period_key, [])
                          if x.get('id') == p['id'] or x.get('term_id') == p['id'] or x.get('semester_id') == p['id']), None)
            row_values.append(_as_float(entry.get('average')) if entry else None)
        row_values += [_as_float(r.get(average_key)), r.get('letter_grade') or '', 'PASS' if r.get('is_passing') else ('FAIL' if r.get('is_passing') is False else '')]

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_i, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            # Percentage columns (every period column + Year Average) get a
            # real numeric value with a % display format, not a string —
            # so the teacher can still sort/filter numerically in Excel.
            if col >= 5 and col <= 4 + n_periods + 1 and isinstance(value, (int, float)):
                cell.number_format = '0.0"%"'

        # Highlight the top 3 ranks, same "medal" idea as the on-screen list.
        rank_val = r.get(rank_key)
        if rank_val in (1, 2, 3):
            medal_fill = {1: 'FFF3C4', 2: 'F0F0F0', 3: 'F5D9B8'}[rank_val]
            for col in range(1, total_cols + 1):
                ws.cell(row=row_i, column=col).fill = PatternFill(start_color=medal_fill, end_color=medal_fill, fill_type='solid')

        row_i += 1

    # Reasonable column widths — name/ID columns wider, numeric columns narrow.
    widths = [7, 8, 26, 16] + [12] * n_periods + [13, 9, 9]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Print setup — landscape + fit-to-width, so a teacher who prints this
    # (rather than just viewing/sorting it on screen) gets every column on
    # one page instead of the columns splitting across pages, which is
    # what plain default portrait page setup does once there are more
    # than ~5-6 columns (verified by rendering this workbook to PDF).
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def xlsx_http_response(data_bytes, filename):
    response = HttpResponse(
        data_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
